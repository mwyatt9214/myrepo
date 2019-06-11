from __future__ import print_function
#import httplib2
import os
#import smtplib
#mport imaplib
import win32com.client as win32
import urllib.request
#from bs4 import BeautifulSoup
import xlwt
import xlrd
import csv
#from itertools import izip
import pandas as pd
import xlsxwriter
from openpyxl import load_workbook
import pyexcel as p
import html5lib
from selenium import webdriver
import requests
from lxml import html
import time
import glob



#from apiclient import discovery
#rom oauth2client import client
#from oauth2client import tools
#from oauth2client.file import Storage

#try:
#   import argparse
#   flags = argparse.ArgumentParser(parents=[tools.argparser]).parse_args()
#except ImportError:
#   flags = None

# If modifying these scopes, delete your previously saved credentials
# at ~/.credentials/sheets.googleapis.com-python-quickstart.json
SCOPES = 'https://www.googleapis.com/auth/spreadsheets.readonly'
CLIENT_SECRET_FILE = 'client_secret.json'
APPLICATION_NAME = 'Google Sheets API Python Quickstart'


def get_credentials():
    """Gets valid user credentials from storage.

    If nothing has been stored, or if the stored credentials are invalid,
    the OAuth2 flow is completed to obtain the new credentials.

    Returns:
        Credentials, the obtained credential.
    """
    home_dir = os.path.expanduser('~')
    credential_dir = os.path.join(home_dir, '.credentials')
    if not os.path.exists(credential_dir):
        os.makedirs(credential_dir)
    credential_path = os.path.join(credential_dir,
                                   'sheets.googleapis.com-python-quickstart.json')

    store = Storage(credential_path)
    credentials = store.get()
    if not credentials or credentials.invalid:
        flow = client.flow_from_clientsecrets(CLIENT_SECRET_FILE, SCOPES)
        flow.user_agent = APPLICATION_NAME
        if flags:
            credentials = tools.run_flow(flow, store, flags)
        else: # Needed only for compatibility with Python 2.6
            credentials = tools.run(flow, store)
        print('Storing credentials to ' + credential_path)
    return credentials

def send_email(user, pwd, recipient, subject, body):
    import smtplib

    gmail_user = user
    gmail_pwd = pwd
    FROM = user
    TO = recipient if type(recipient) is list else [recipient]
    SUBJECT = subject
    TEXT = body

    # Prepare actual message
    message = """From: %s\nTo: %s\nSubject: %s\n\n%s
    """ % (FROM, ", ".join(TO), SUBJECT, TEXT)
    #try:
    server = smtplib.SMTP("smtp.gmail.com", 587)
    server.ehlo()
    server.starttls()
    server.login(gmail_user, gmail_pwd)
    server.sendmail(FROM, TO, message)
    server.close()
    print ('successfully sent the mail')
    #except:
        #print ("failed to send mail")
    return "done"
def send_email_outlook(recipient,subject,body):
    import win32com.client as win32
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = recipient
    mail.Subject = subject
    mail.Body = body
    mail.HTMLBody = '<h2>'+ body +'</h2>'# this field is optional

#In case you want to attach a file to the email
    #attachment  = "Path to the attachment"
    #mail.Attachments.Add(attachment)

    mail.Send()
def get_gmail():
   # try:
    mail = imaplib.IMAP4_SSL("smtp.gmail.com")
    mail.login("mwyatt9214@gmail.com","CJ15035300")
    mail.select('inbox')

    typ, data = mail.search(None, 'ALL')
    for num in range(2):
        typ, data = mail.fetch(num, '(RFC822)')
        print ('Message %s\n%s\n' % (num, data[0][1]))
    mail.close()

   # except Exception:
       # e = Exception
       # print (str(e))

def download_csv(Pos):
        global final_file
        if Pos == 'QB':
            link = 'https://www.fantasypros.com/nfl/stats/qb.php'
            final_file = "FantasyPros_Fantasy_Football_Statistics_QB.csv"
        elif Pos ==  'RB':
            link = 'https://www.fantasypros.com/nfl/stats/rb.php'
            final_file = "FantasyPros_Fantasy_Football_Statistics_RB.csv"
        elif Pos == 'WR':
            link = 'https://www.fantasypros.com/nfl/stats/wr.php'
            final_file = "FantasyPros_Fantasy_Football_Statistics_WR.csv"
        elif Pos == 'TE':
            link = 'https://www.fantasypros.com/nfl/stats/te.php'
            final_file = "FantasyPros_Fantasy_Football_Statistics_TE.csv"
        elif Pos == 'Snap': 
            link = 'https://www.fantasypros.com/nfl/reports/snap-counts/'
            final_file = "FantasyPros_Fantasy_Football_2017_Offense_Snap_Counts.csv"
        elif Pos == 'Target':
            link = 'https://www.fantasypros.com/nfl/reports/targets/'
            final_file = "FantasyPros_Fantasy_Football_2017_Target_Leaders.csv"
        
        #session_requests = requests.session()


        url = "https://www.fantasypros.com/nfl/stats/qb.php?ownership=e&scoring=HALF"
        #login_payload = {"username":"<mwyatt9214@gmail.com>","password":"<CJ15035300>","crsfmiddlewaretoken":"<d1if2O0wT5qlXGcXnwCCEdzdDWfZv9GB>"}
        #result = session_requests.get(url)
        #tree = html.formstring(result.text)
        #autheticity_token = list(set(tree.xpath("//input[@name='csrfmiddlewaretoken']/@value")))[0]

        #result = session_requests.post(url, data = login_payload, headers = dict(referer=url))


        #login_payload = {"username":"<mwyatt9214@gmail.com>","password":"<CJ15035300>","crsfmiddlewaretoken":"<d1if2O0wT5qlXGcXnwCCEdzdDWfZv9GB>"}

        driver = webdriver.Chrome("C:\\Users\Mateo\chromedriver.exe")
        driver.implicitly_wait(10)
        driver.get('https://secure.fantasypros.com/accounts/login/?next=http://www.fantasypros.com/index.php?loggedin&loggedout')
        time.sleep(3)
        username=driver.find_element_by_id('id_username')
        password=driver.find_element_by_id('id_password')
        time.sleep(3)
        username.send_keys("mrw7fq@virginia.edu")
        password.send_keys("CJ15035300")

        driver.find_element_by_xpath('/html/body/div/div[2]/div/div[2]/form/button').click()
        #print(driver.title)

        driver.get(link)
        driver.find_element_by_xpath('/html/body/div[2]/div[6]/div/div[1]/div/div[1]/div[2]/a[2]').click()
        time.sleep(3)
        '''list_of_files = glob.glob('\Downloads\*') # * means all if need specific format then *.csv
        latest_file = max(list_of_files, key=os.path.getctime)
        print (latest_file)
            '''


        '''content = urllib.request.urlopen(url)
        content1 = content.read()
        soup = BeautifulSoup(content,'lxml')
        print(soup.get_text())
        for link in soup.find_all("a"):
            print(link.get("href"))
            print("im here")
            '''
        '''
        for link in soup.find_all('a'):
            print(link.get('href'))
        resp = urllib.request.urlopen(link)
        res = resp.read()
        with open( Pos+".txt", "wb" ) as code :
          code.write( res )
        i = 0
        c=0
        f = open(Pos+'.txt', 'r+')
        row_list = []    
        for row in f:
            row_list.append(row.split('","'))
        for row in row_list:
            row[0] = row[0][1::]
            if Pos == "Target" or Pos =="Snap" or c==0:
                row[-1]=row[-1][0:-2]
            
            else:
                row[-1]=row[-1][0:-3]
            c+=1



                

        #file  = "Fantasy_Football_Analysis.xlsx"
        workbook = xlwt.Workbook()
        worksheet = workbook.add_sheet('Sheet1')
        global a
        global b
        b=0
        a=0
        i = 0 

        for column in row_list:
            for item in range(len(column)):
                try:
                    if column[item].isnumeric() or column[item][-2]==".":
                    #column[item].isalpha() or column[item].count("/")>0 or column[item].count("(")>0 or column[item].count(")")>0:
                        b=1
                        worksheet.write(item, i, float(column[item]))
                        a=1
                    elif column[item].count(",")>0:
                        b=2
                        worksheet.write(item, i, int(column[item].replace(',', '')))
                        a=2
                    elif column[item].count("%")>0:
                        b=3
                        worksheet.write(item, i, int(column[item].replace('%', '')))
                        a=4
                    else:
                        b=4
                        worksheet.write(item, i, column[item])
                        a=3
                except Exception as ex:
                    if len(column[item]) == 1 and i<256:
                        worksheet.write(item, i, column[item])
                    if Pos == "QB"and i<256:
                        print(b, column[item],ex)

                    pass
                if a == 0 and i <256:
                    #print(item, i, b, column[item])
                    worksheet.write(item, i, True)

            workbook.save(Pos+'.xls')
            i+=1
            a=0
            '''
        #owd = os.getcwd()
        #os.chdir(os.path.dirname("\Downloads"))
        p.save_book_as(file_name="C:\\Users\mateo\Downloads\{0}".format(final_file),
               dest_file_name=Pos+'.xlsx')
        
        xlsxFile = Pos+".xlsx"
        wb1 = load_workbook(xlsxFile)
        #os.chdir(owd)
        wb2 = load_workbook('Fantasy_Football_Analysis.xlsx')
        print (list(wb1.get_sheet_names()))

        sheet1 = wb1.get_sheet_by_name(final_file)
        sheet2 = wb2.get_sheet_by_name('{0}_raw'.format(Pos))

        for i,row in enumerate(sheet1.iter_rows()):
            for j,col in enumerate(row):
                sheet2.cell(row=i+1,column=j+1).value = col.value
        wb1.save(xlsxFile)
        wb2.save('Fantasy_Football_Analysis.xlsx')
        
        os.remove("C:\\Users\mateo\Downloads\{0}".format(final_file))
        os.remove(Pos+'.xlsx')

def main():
    """Shows basic usage of the Sheets API.

    Creates a Sheets API service object and prints the names and majors of
    students in a sample spreadsheet:
    https://docs.google.com/spreadsheets/d/1BxiMVs0XRA5nFMdKvBdBZjgmUUqptlbs74OgvE2upms/edit
    """
    """credentials = get_credentials()
    http = credentials.authorize(httplib2.Http())
    discoveryUrl = ('https://sheets.googleapis.com/$discovery/rest?'
                    'version=v4')
    service = discovery.build('sheets', 'v4', http=http,
                              discoveryServiceUrl=discoveryUrl)

    spreadsheetId = '1-DaZAVffa2jSI6-yteGBIchXjljY16wGFXVTdAXYANg'
    rangeName = 'Sheet1!A1:E'
    result = service.spreadsheets().values().get(
        spreadsheetId=spreadsheetId, range=rangeName).execute()
    values = result.get("values")
    print(values)
    if not values:
        print('No data found.')
    else:
        print('Name, Major:')
        for row in values:
            # Print columns A and E, which correspond to indices 0 and 4.
            print('%s, %s' % (row[0], row[1]))
    #send_email("mwyatt9214@gmail.com","CJ15035300","mwyatt9214@gmail.com","testPY",values)\
    #get_gmail()
    send_email_outlook("mwyatt@tesla.com","Test", 'Testing123, automated outlook email')
        """
    download_csv("QB")
    print("QB done")
    download_csv('RB')
    print("RB done")
    download_csv('WR')
    print("WR done")
    download_csv('TE')
    print("TE done")
    download_csv('Snap')
    print("Snap done")
    download_csv('Target')
    print("Target done")
        



if __name__ == '__main__':
    main()
    